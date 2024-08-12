from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ایجاد یک کتاب کار جدید
wb = Workbook()
ws = wb.active
ws.title = "فیش حقوقی"

# تنظیمات عمومی
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
right_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# سربرگ
ws.merge_cells('A1:I1')
ws['A1'] = "جاوید ساخت | همیار تخصصی اکسل"
ws['A1'].font = bold_font
ws['A1'].alignment = center_alignment
ws['A1'].border = thin_border

ws.merge_cells('A2:I2')
ws['A2'] = "فیش حقوق"
ws['A2'].font = bold_font
ws['A2'].alignment = center_alignment

ws.merge_cells('A3:I3')
ws['A3'] = "فیش حقوقی اردیبهشت سال 1399"
ws['A3'].alignment = center_alignment
ws['A3'].border = thin_border

# اطلاعات شخصی و شرکتی
info_labels = [
    ("نام:", "مسعود آزادی"),
    ("کد ملی:", "1234567890"),
    ("شماره شناسنامه:", "878787"),
    ("وضعیت تاهل:", "متاهل"),
    ("محل خدمت:", "حسابداری"),
    ("عنوان شغلی:", "حسابدار ارشد"),
    ("شماره پرسنلی:", "3030"),
    ("شماره حساب:", "45345345"),
    ("شماره بیمه:", "567834")
]

# تنظیم اطلاعات شخصی و شرکتی
start_row = 4
start_col = 9
for i, (label, value) in enumerate(info_labels):
    row = start_row + (i % 3)
    col = start_col - (i // 3) * 2
    ws.cell(row=row, column=col, value=label)
    ws.cell(row=row, column=col + 1, value=value)

# تنظیم کادر برای اطلاعات شخصی و شرکتی
for row in ws.iter_rows(min_row=4, max_row=6, min_col=7, max_col=9):
    for cell in row:
        cell.border = thin_border
        cell.alignment = right_alignment

# اطلاعات کارکرد
ws.merge_cells('A8:B8')
ws['A8'] = "اطلاعات کارکرد"
ws['A8'].alignment = center_alignment
ws['A8'].border = thin_border

work_labels = [
    ("کارکرد روزانه", "22"),
    ("ساعت اضافه کار", "15"),
    ("جمعه کاری", "2")
]

for i, (label, value) in enumerate(work_labels):
    ws.cell(row=9 + i, column=1, value=label)
    ws.cell(row=9 + i, column=2, value=value)

# شرح پرداخت و کسورات
ws.merge_cells('D8:F8')
ws['D8'] = "شرح پرداخت"
ws['D8'].alignment = center_alignment
ws['D8'].border = thin_border

payment_details = [
    ("حقوق پایه", "11315400"),
    ("پایه سنوات", "1150000"),
    ("حق مسکن", "1000000"),
    ("حق اولاد", "600000"),
    ("جمع اضافه کاری", "2332400"),
    ("سایر مزایا", "1500000")
]

for i, (description, amount) in enumerate(payment_details):
    ws.cell(row=9 + i, column=4, value=description)
    ws.cell(row=9 + i, column=5, value=amount)
    ws.cell(row=9 + i, column=6, value="ریال")

ws.merge_cells('G8:H8')
ws['G8'] = "کسورات"
ws['G8'].alignment = center_alignment
ws['G8'].border = thin_border

deductions = [
    ("بیمه", "1513243"),
    ("مالیات", "1245352"),
    ("سایر کسورات", "500000")
]

for i, (description, amount) in enumerate(deductions):
    ws.cell(row=9 + i, column=7, value=description)
    ws.cell(row=9 + i, column=8, value=amount)

# تنظیم کادر برای اطلاعات کارکرد، شرح پرداخت و کسورات
for row in ws.iter_rows(min_row=8, max_row=14, min_col=1, max_col=9):
    for cell in row:
        cell.border = thin_border
        cell.alignment = right_alignment

# جمع پرداختی و خالص دریافتی
ws['D16'] = "جمع کل مزایا:"
ws['E16'] = "=SUM(E9:E14)"
ws['F16'] = "ریال"

ws['G16'] = "جمع کل کسورات:"
ws['H16'] = "=SUM(H9:H11)"
ws['H16'].alignment = right_alignment

ws.merge_cells('A18:B18')
ws['A18'] = "=E16-H16"
ws['A18'].font = Font(bold=True)
ws['C18'] = "ریال"

# تنظیم کادر برای جمع پرداختی و خالص دریافتی
for row in ws.iter_rows(min_row=16, max_row=18, min_col=1, max_col=9):
    for cell in row:
        cell.border = thin_border
        cell.alignment = right_alignment

# تنظیم عرض ستون‌ها برای نمایش بهتر
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
for col in columns:
    cell_lengths = [len(str(cell.value)) for cell in ws[col] if cell.value]
    if cell_lengths:
        max_length = max(cell_lengths)
    else:
        max_length = 10  # مقدار پیش‌فرض در صورت خالی بودن ستون
    adjusted_width = max_length + 2
    ws.column_dimensions[col].width = adjusted_width

# ذخیره فایل اکسل
wb.save("fish_hoghooghi_final.xlsx")

print("فیش حقوقی با موفقیت ایجاد شد و در فایل 'fish_hoghooghi_final.xlsx' ذخیره شد.")
